import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response, g, send_file
from sqlalchemy.exc import IntegrityError, OperationalError
from sqlalchemy import func, distinct
from waitress import serve
from datetime import datetime, time, timedelta
from dateutil.relativedelta import relativedelta
import uuid
import time as time_module
from threading import Thread, Lock
import socket
from werkzeug.security import generate_password_hash, check_password_hash
import secrets
from functools import wraps
from flask_socketio import SocketIO, emit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

from database import db, Config, MonitoredGroup, Keyword, MatchedMessage, DB_URI, User, Session, auto_upgrade_database
from telegram_monitor import start_monitoring, stop_monitoring, is_running, keyword_automatons, automatons_lock
from telegram_utils import get_group_details, get_my_groups, batch_join_groups

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = DB_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', 'your_very_secret_key_here_please_change_me')

# 初始化 SocketIO
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

batch_join_tasks = {}
tasks_lock = Lock()

db.init_app(app)

with app.app_context():
    db.create_all()
    # 自动检查并升级数据库结构
    auto_upgrade_database()


# 用于检查用户会话，并实现60分钟过期和自动续期
def check_session_and_renew():
    session_id = request.cookies.get('session_id')
    if not session_id:
        return None

    user_session = Session.query.filter_by(id=session_id).first()
    if not user_session:
        return None

    # 检查会话是否过期
    if user_session.expiration_time < datetime.now():
        db.session.delete(user_session)
        db.session.commit()
        return None

    # 会话未过期，自动续期60分钟
    user_session.expiration_time = datetime.now() + timedelta(minutes=60)
    db.session.commit()

    return db.session.get(User, user_session.user_id)

@app.before_request
def load_logged_in_user():
    g.user = check_session_and_renew()

def check_config_exists():
    """检查 Telegram 配置是否存在"""
    config = Config.query.first()
    return config is not None

# 免配置检查的路由
NO_CONFIG_ALLOWED = {'setup', 'static', 'login', 'logout', 'verify', 'verify_status'}

@app.before_request
def redirect_to_setup_if_no_config():
    """如果未配置 Telegram，重定向到配置向导（除非是允许的路由）"""
    if request.endpoint in NO_CONFIG_ALLOWED:
        return None
    try:
        if not check_config_exists():
            return redirect(url_for('setup'))
    except Exception:
        # 数据库可能还没准备好，允许访问 setup
        if request.endpoint != 'setup':
            return redirect(url_for('setup'))

def login_required(f):
    @wraps(f) # 保留原函数的元信息
    def decorated_function(*args, **kwargs):
        if g.user is None:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/setup', methods=['GET', 'POST'])
def setup():
    """Telegram API 配置向导（网页端）"""
    config = Config.query.first()
    if config:
        # 已有配置，跳转到登录页
        return redirect(url_for('login'))

    if request.method == 'POST':
        api_id = request.form.get('api_id', '').strip()
        api_hash = request.form.get('api_hash', '').strip()
        phone_number = request.form.get('phone_number', '').strip()

        # 验证
        if not api_id.isdigit():
            flash('API ID 必须为纯数字', 'danger')
            return render_template('setup.html')
        if len(api_hash) < 10:
            flash('API Hash 格式不正确', 'danger')
            return render_template('setup.html')
        if not phone_number.startswith('+'):
            flash('手机号码必须以 + 开头（含国际区号）', 'danger')
            return render_template('setup.html')

        # 保存配置
        new_config = Config(
            api_id=api_id,
            api_hash=api_hash,
            phone_number=phone_number
        )
        db.session.add(new_config)
        db.session.commit()

        flash('Telegram 配置已保存！正在启动监控服务...', 'success')

        # 异步启动 Telegram 监控
        def start_monitoring_async():
            try:
                import telegram_monitor
                telegram_monitor.websocket_broadcast_callback = broadcast_new_message
                start_monitoring()
                print("[异步] Telegram 监控服务已启动")
            except Exception as e:
                print(f"[异步] 监控启动失败: {e}")

        thread = Thread(target=start_monitoring_async, daemon=True)
        thread.start()

        return redirect(url_for('verify'))

    return render_template('setup.html')

@app.route('/verify', methods=['GET', 'POST'])
def verify():
    """Telegram 验证码输入页面"""
    from telegram_monitor import verification_manager, is_running as tg_running

    # 已经验证通过，跳转登录
    if tg_running or verification_manager.step == 'done':
        return redirect(url_for('login'))

    if request.method == 'POST':
        step = request.form.get('step', 'code')
        value = request.form.get('code' if step == 'code' else 'password', '').strip()
        if value:
            verification_manager.submit_code(value)
            flash('验证码已提交，正在验证...', 'info')
        return redirect(url_for('verify'))

    # 获取手机号用于显示（隐藏中间几位）
    config = Config.query.first()
    phone_display = ''
    if config and config.phone_number:
        p = config.phone_number
        if len(p) > 7:
            phone_display = p[:3] + '****' + p[-3:]
        else:
            phone_display = p

    return render_template('verify.html', phone=phone_display)

@app.route('/verify/status')
def verify_status():
    """返回验证状态（AJAX 轮询）"""
    from telegram_monitor import verification_manager, is_running as tg_running
    if tg_running:
        return jsonify({'step': 'done', 'message': 'Telegram 已连接成功！'})
    status = verification_manager.get_status()
    return jsonify(status)

@app.route('/')
@login_required
def index():
    return redirect(url_for('dashboard')) # 重定向到仪表盘

@app.route('/dashboard')
@login_required
def dashboard():
    """数据可视化仪表盘首页"""
    return render_template('dashboard.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if g.user: # 如果用户已经登录，则重定向到主页
        return redirect(url_for('groups')) # 更改默认重定向到群组管理页面

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()

        if user and check_password_hash(user.password_hash, password):
            # 登录成功，创建会话
            session_id = str(uuid.uuid4())
            expiration_time = datetime.now() + timedelta(minutes=60)
            new_session = Session(id=session_id, user_id=user.id, expiration_time=expiration_time)
            db.session.add(new_session)
            db.session.commit()

            response = make_response(redirect(request.args.get('next') or url_for('groups')))
            response.set_cookie('session_id', session_id, httponly=True, expires=expiration_time)
            flash('登录成功！', 'success')
            return response
        else:
            flash('用户名或密码错误。' , 'danger')
            return redirect(url_for('login', next=request.args.get('next')))
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    session_id = request.cookies.get('session_id')
    if session_id:
        user_session = Session.query.filter_by(id=session_id).first()
        if user_session:
            db.session.delete(user_session)
            db.session.commit()
    
    response = make_response(redirect(url_for('login')))
    response.set_cookie('session_id', '', expires=0) # 清除cookie
    flash('您已成功注销。' , 'info')
    return response


@app.route('/api/dashboard/stats')
@login_required
def dashboard_stats():
    """获取仪表盘统计数据"""
    try:
        now = datetime.now()
        today_start = datetime.combine(now.date(), time.min)
        week_start = today_start - timedelta(days=now.weekday())
        month_start = datetime(now.year, now.month, 1)
        
        # 今日统计
        today_count = MatchedMessage.query.filter(
            MatchedMessage.message_date >= today_start
        ).count()
        
        # 本周统计
        week_count = MatchedMessage.query.filter(
            MatchedMessage.message_date >= week_start
        ).count()
        
        # 本月统计
        month_count = MatchedMessage.query.filter(
            MatchedMessage.message_date >= month_start
        ).count()
        
        # 总计
        total_count = MatchedMessage.query.count()
        
        # 活跃群组数（处理 None 情况）
        active_groups = db.session.query(func.count(distinct(MatchedMessage.group_name))).scalar() or 0
        
        # 总群组数
        total_groups = MonitoredGroup.query.count()
        
        # 计算环比（与上一周期对比）- 优化逻辑
        yesterday_start = today_start - timedelta(days=1)
        yesterday_count = MatchedMessage.query.filter(
            MatchedMessage.message_date >= yesterday_start,
            MatchedMessage.message_date < today_start
        ).count()
        
        if yesterday_count > 0:
            today_change = ((today_count - yesterday_count) / yesterday_count * 100)
        elif today_count > 0:
            today_change = 999  # 表示新增（前端显示为 "新增"）
        else:
            today_change = 0
        
        last_week_start = week_start - timedelta(days=7)
        last_week_count = MatchedMessage.query.filter(
            MatchedMessage.message_date >= last_week_start,
            MatchedMessage.message_date < week_start
        ).count()
        
        if last_week_count > 0:
            week_change = ((week_count - last_week_count) / last_week_count * 100)
        elif week_count > 0:
            week_change = 999
        else:
            week_change = 0
        
        last_month_start = month_start - relativedelta(months=1)
        last_month_count = MatchedMessage.query.filter(
            MatchedMessage.message_date >= last_month_start,
            MatchedMessage.message_date < month_start
        ).count()
        
        if last_month_count > 0:
            month_change = ((month_count - last_month_count) / last_month_count * 100)
        elif month_count > 0:
            month_change = 999
        else:
            month_change = 0
        
        return jsonify({
            'today': {
                'count': today_count,
                'change': round(today_change, 1)
            },
            'week': {
                'count': week_count,
                'change': round(week_change, 1)
            },
            'month': {
                'count': month_count,
                'change': round(month_change, 1)
            },
            'total': total_count,
            'active_groups': active_groups,
            'total_groups': total_groups
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/hot_keywords')
@login_required
def dashboard_hot_keywords():
    """获取热词统计"""
    try:
        period = request.args.get('period', 'week')  # today, week, month, all
        limit = request.args.get('limit', 50, type=int)
        
        # 数据验证
        if period not in ['today', 'week', 'month', 'all']:
            period = 'week'
        if limit < 1 or limit > 100:
            limit = 50
        
        now = datetime.now()
        if period == 'today':
            start_date = datetime.combine(now.date(), time.min)
        elif period == 'week':
            start_date = datetime.combine(now.date(), time.min) - timedelta(days=now.weekday())
        elif period == 'month':
            start_date = datetime(now.year, now.month, 1)
        else:  # all
            start_date = datetime(2000, 1, 1)
        
        # 统计关键词频率
        keyword_stats = db.session.query(
            MatchedMessage.matched_keyword,
            func.count(MatchedMessage.matched_keyword).label('count')
        ).filter(
            MatchedMessage.message_date >= start_date
        ).group_by(
            MatchedMessage.matched_keyword
        ).order_by(
            func.count(MatchedMessage.matched_keyword).desc()
        ).limit(limit).all()
        
        result = [{'keyword': kw, 'count': cnt} for kw, cnt in keyword_stats]
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/group_activity')
@login_required
def dashboard_group_activity():
    """获取群组活跃度统计"""
    try:
        period = request.args.get('period', 'week')
        limit = request.args.get('limit', 10, type=int)
        
        # 数据验证
        if period not in ['today', 'week', 'month', 'all']:
            period = 'week'
        if limit < 1 or limit > 50:
            limit = 10
        
        now = datetime.now()
        if period == 'today':
            start_date = datetime.combine(now.date(), time.min)
        elif period == 'week':
            start_date = datetime.combine(now.date(), time.min) - timedelta(days=now.weekday())
        elif period == 'month':
            start_date = datetime(now.year, now.month, 1)
        else:
            start_date = datetime(2000, 1, 1)
        
        # 统计群组活跃度
        group_stats = db.session.query(
            MatchedMessage.group_name,
            func.count(MatchedMessage.id).label('count')
        ).filter(
            MatchedMessage.message_date >= start_date
        ).group_by(
            MatchedMessage.group_name
        ).order_by(
            func.count(MatchedMessage.id).desc()
        ).limit(limit).all()
        
        result = [{'group_name': gn, 'count': cnt} for gn, cnt in group_stats]
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/trends')
@login_required
def dashboard_trends():
    """获取匹配趋势数据"""
    try:
        period = request.args.get('period', '7d')  # 7d, 30d, 12m
        
        # 数据验证
        if period not in ['7d', '30d', '12m']:
            period = '7d'
        
        now = datetime.now()
        
        if period == '7d':
            # 最近7天，按天统计（优化：一次查询）
            start_date = datetime.combine(now.date(), time.min) - timedelta(days=6)
            
            # 一次性查询所有数据并按日期分组
            results = db.session.query(
                func.date(MatchedMessage.message_date).label('date'),
                func.count(MatchedMessage.id).label('count')
            ).filter(
                MatchedMessage.message_date >= start_date
            ).group_by(
                func.date(MatchedMessage.message_date)
            ).all()
            
            # 创建日期到数量的映射
            date_count_map = {str(r.date): r.count for r in results}
            
            # 填充所有日期（包括没有数据的日期）
            trends = []
            for i in range(7):
                day = start_date + timedelta(days=i)
                date_str = day.strftime('%Y-%m-%d')
                trends.append({
                    'date': date_str,
                    'label': day.strftime('%m/%d'),
                    'count': date_count_map.get(date_str, 0)
                })
                
        elif period == '30d':
            # 最近30天，按天统计（优化：一次查询）
            start_date = datetime.combine(now.date(), time.min) - timedelta(days=29)
            
            results = db.session.query(
                func.date(MatchedMessage.message_date).label('date'),
                func.count(MatchedMessage.id).label('count')
            ).filter(
                MatchedMessage.message_date >= start_date
            ).group_by(
                func.date(MatchedMessage.message_date)
            ).all()
            
            date_count_map = {str(r.date): r.count for r in results}
            
            trends = []
            for i in range(30):
                day = start_date + timedelta(days=i)
                date_str = day.strftime('%Y-%m-%d')
                trends.append({
                    'date': date_str,
                    'label': day.strftime('%m/%d'),
                    'count': date_count_map.get(date_str, 0)
                })
        else:  # 12m
            # 最近12个月，按月统计（优化：一次查询）
            current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            start_month = current_month_start - relativedelta(months=11)
            
            # 一次性查询所有数据并按月分组
            results = db.session.query(
                func.date_format(MatchedMessage.message_date, '%Y-%m').label('month'),
                func.count(MatchedMessage.id).label('count')
            ).filter(
                MatchedMessage.message_date >= start_month
            ).group_by(
                func.date_format(MatchedMessage.message_date, '%Y-%m')
            ).all()
            
            month_count_map = {r.month: r.count for r in results}
            
            trends = []
            for i in range(12):
                month_start = current_month_start - relativedelta(months=11-i)
                month_str = month_start.strftime('%Y-%m')
                trends.append({
                    'date': month_str,
                    'label': month_start.strftime('%Y/%m'),
                    'count': month_count_map.get(month_str, 0)
                })
        
        return jsonify(trends)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/config', methods=['GET', 'POST'])
@login_required # 添加鉴权装饰器
def config():
    config_item = Config.query.first()
    if request.method == 'POST':
        api_id = request.form.get('api_id')
        api_hash = request.form.get('api_hash')
        phone_number = request.form.get('phone_number')
        
        # 通知配置
        notification_type = request.form.get('notification_type', 'none')
        dingtalk_webhook = request.form.get('dingtalk_webhook')
        dingtalk_secret = request.form.get('dingtalk_secret')
        wecom_webhook = request.form.get('wecom_webhook')

        if config_item:
            config_item.api_id = api_id
            config_item.api_hash = api_hash
            config_item.phone_number = phone_number
            config_item.notification_type = notification_type
            config_item.dingtalk_webhook = dingtalk_webhook
            config_item.dingtalk_secret = dingtalk_secret
            config_item.wecom_webhook = wecom_webhook
        else:
            config_item = Config(
                api_id=api_id,
                api_hash=api_hash,
                phone_number=phone_number,
                notification_type=notification_type,
                dingtalk_webhook=dingtalk_webhook,
                dingtalk_secret=dingtalk_secret,
                wecom_webhook=wecom_webhook
            )
            db.session.add(config_item)
        
        db.session.commit()
        flash('配置已成功保存！', 'success')
        return redirect(url_for('config'))

    if not config_item:
        config_item = {
            'api_id': '', 
            'api_hash': '', 
            'phone_number': '',
            'notification_type': 'none',
            'dingtalk_webhook': '',
            'dingtalk_secret': '',
            'wecom_webhook': ''
        }

    return render_template('config.html', config=config_item, is_running=is_running)

@app.route('/api/batch_join', methods=['POST'])
@login_required # 添加鉴权装饰器
def start_batch_join():
    data = request.get_json()
    links_text = data.get('links', '')
    delay = data.get('delay', 60)

    links = [link.strip() for link in links_text.splitlines() if link.strip()]
    if not links:
        return jsonify({'error': '链接列表不能为空。'}), 400
    try:
        delay = int(delay)
        if delay < 20:
            delay = 20
    except (ValueError, TypeError):
        delay = 60

    task_id = str(uuid.uuid4())
    
    with tasks_lock:
        batch_join_tasks[task_id] = {
            'status': 'pending',
            'log': ['[INFO] 任务已创建，正在等待后台线程启动...'],
            'stop_requested': False,
            'total': len(links),
            'current': 0
        }

    thread = Thread(target=batch_join_groups, args=(task_id, links, delay, batch_join_tasks, tasks_lock))
    thread.daemon = True
    thread.start()

    return jsonify({'task_id': task_id})

@app.route('/api/batch_join/status/<task_id>', methods=['GET'])
@login_required # 添加鉴权装饰器
def get_batch_join_status(task_id):
    with tasks_lock:
        task = batch_join_tasks.get(task_id)
    
    if not task:
        return jsonify({'error': '任务未找到'}), 404
    
    return jsonify(task)

@app.route('/api/batch_join/stop/<task_id>', methods=['POST'])
@login_required # 添加鉴权装饰器
def stop_batch_join(task_id):
    with tasks_lock:
        task = batch_join_tasks.get(task_id)
        if task and task['status'] == 'running':
            task['stop_requested'] = True
            task['log'].append('[INFO] 收到停止请求，将在当前操作完成后中止...')
            return jsonify({'message': '停止请求已发送。'})

    return jsonify({'error': '任务未找到或已结束'}), 404


@app.route('/status')
@login_required # 添加鉴权装饰器
def status():
    from telegram_monitor import client_thread
    is_alive = client_thread is not None and client_thread.is_alive()
    return jsonify({'is_running': is_alive})

@app.route('/control/test_dingtalk', methods=['POST'])
@login_required # 添加鉴权装饰器
def test_dingtalk():
    config = Config.query.first()
    if not config or not config.dingtalk_webhook:
        flash('请先保存钉钉Webhook地址。', 'warning')
        return redirect(url_for('config'))

    title = "测试消息"
    message = "这是一条来自Telegram监控系统的测试消息。"
    
    from telegram_monitor import send_to_dingtalk
    result = send_to_dingtalk(config.dingtalk_webhook, config.dingtalk_secret, title, message, is_test=True)
    
    flash(f'钉钉测试结果: {result}', 'info')
    return redirect(url_for('config'))

@app.route('/control/test_wecom', methods=['POST'])
@login_required # 添加鉴权装饰器
def test_wecom():
    config = Config.query.first()
    if not config or not config.wecom_webhook:
        flash('请先保存企业微信Webhook地址。', 'warning')
        return redirect(url_for('config'))

    title = "测试消息"
    message = "这是一条来自Telegram监控系统的测试消息。\n\n> **测试项目**: 企业微信机器人\n> **测试时间**: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    from telegram_monitor import send_to_wecom
    result = send_to_wecom(config.wecom_webhook, title, message, is_test=True)
    
    flash(f'企业微信测试结果: {result}', 'info')
    return redirect(url_for('config'))


@app.route('/groups', methods=['GET', 'POST'])
@login_required # 添加鉴权装饰器
def groups():
    if request.method == 'POST':
        group_identifier_input = request.form.get('group_identifier')
        
        if group_identifier_input:
            try:
                details = get_group_details(group_identifier_input)

                if details.get('error'):
                    flash(f"添加失败: {details['error']}", 'danger')
                else:
                    new_group = MonitoredGroup(
                        group_identifier=details['identifier'], 
                        group_name=details['name'],
                        logo_path=details['logo_path']
                    )
                    db.session.add(new_group)
                    try:
                        db.session.commit()
                        flash(f"群组 '{details['name']}' 添加成功！", 'success')
                    except IntegrityError:
                        db.session.rollback()
                        flash(f"群组 '{details['name']}' 已经存在。", 'danger')
            except OperationalError as e:
                db.session.rollback()
                if "database is locked" in str(e).lower():
                    flash("系统正忙，请稍后重试。可能是后台正在进行Telegram操作。", 'warning')
                else:
                    flash(f"发生数据库错误: {e}", 'danger')
            except Exception as e:
                db.session.rollback()
                flash(f"发生未知错误: {e}", 'danger')
        
        return redirect(url_for('groups'))
    
    # 处理GET请求
    search_query = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)  # 获取当前页码，默认第1页
    per_page = 50  # 每页显示50条
    
    query = MonitoredGroup.query
    if search_query:
        query = query.filter(MonitoredGroup.group_name.ilike(f'%{search_query}%'))
    
    # 使用分页
    pagination = query.order_by(MonitoredGroup.group_name).paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    
    return render_template(
        'groups.html', 
        groups=pagination.items,  # 当前页的群组
        pagination=pagination,     # 分页对象
        search_query=search_query
    )

@app.route('/add_my_groups', methods=['GET'])
@login_required # 添加鉴权装饰器
def add_my_groups_page():
    return render_template('add_my_groups.html')


@app.route('/api/get_my_groups', methods=['GET'])
@login_required # 添加鉴权装饰器
def api_get_my_groups():
    try:
        monitored_ids = {g.group_identifier for g in MonitoredGroup.query.all()}
    finally:
        db.session.remove()

    result = get_my_groups()
    if result.get('error'):
        return jsonify({'error': result['error']}), 500
    
    my_groups = [g for g in result['groups'] if g['id'] not in monitored_ids]
    
    return jsonify({'groups': my_groups})


@app.route('/groups/batch_add', methods=['POST'])
@login_required # 添加鉴权装饰器
def batch_add_groups():
    groups_to_add = request.form.getlist('groups')
    added_count = 0
    skipped_count = 0
    for group_data in groups_to_add:
        parts = group_data.split('|||')
        if len(parts) != 3: continue

        group_id, group_name, logo_path = parts
        
        exists = MonitoredGroup.query.filter_by(group_identifier=group_id).first()
        if not exists:
            new_group = MonitoredGroup(
                group_identifier=group_id,
                group_name=group_name,
                logo_path=logo_path if logo_path != 'None' else None
            )
            db.session.add(new_group)
            added_count += 1
        else:
            skipped_count += 1
    
    if added_count > 0:
        db.session.commit()
        flash(f'成功添加 {added_count} 个新群组！', 'success')
    if skipped_count > 0:
        flash(f'跳过 {skipped_count} 个已存在的群组。', 'info')

    return redirect(url_for('groups'))


@app.route('/groups/delete/<int:group_id>')
@login_required # 添加鉴权装饰器
def delete_group(group_id):
    group_to_delete = MonitoredGroup.query.get_or_404(group_id)
    db.session.delete(group_to_delete)
    db.session.commit()
    flash('群组已删除。', 'info')
    return redirect(url_for('groups'))

@app.route('/groups/batch_delete', methods=['POST'])
@login_required
def batch_delete_groups():
    group_ids = request.form.getlist('group_ids')
    if not group_ids:
        flash('没有选择任何群组。', 'warning')
        return redirect(url_for('groups'))

    groups_to_delete = MonitoredGroup.query.filter(MonitoredGroup.id.in_(group_ids)).all()
    
    deleted_count = len(groups_to_delete)
    for group in groups_to_delete:
        db.session.delete(group)

    db.session.commit()
    flash(f'成功删除 {deleted_count} 个群组!', 'success')
    return redirect(url_for('groups'))

@app.route('/keywords', methods=['GET', 'POST'])
@login_required # 添加鉴权装饰器
def keywords():
    if request.method == 'POST':
        keywords_text = request.form.get('keywords_text', '').strip()
        group_ids = request.form.getlist('groups')

        if not keywords_text:
            flash('关键词列表不能为空。', 'danger')
        elif not group_ids:
            flash('必须至少选择一个群组。', 'danger')
        else:
            keywords_list = [kw.strip() for kw in keywords_text.splitlines() if kw.strip()]
            added_count = 0
            skipped_count = 0
            groups = MonitoredGroup.query.filter(MonitoredGroup.id.in_(group_ids)).all()

            for keyword_text in keywords_list:
                existing_keyword = Keyword.query.filter_by(text=keyword_text).first()
                if existing_keyword:
                    skipped_count += 1
                else:
                    new_keyword = Keyword(text=keyword_text)
                    new_keyword.groups.extend(groups)
                    db.session.add(new_keyword)
                    added_count += 1
            
            if added_count > 0:
                db.session.commit()
                
                # 性能优化: 清空相关群组的AC自动机缓存
                with automatons_lock:
                    for group in groups:
                        keyword_automatons.pop(group.id, None)
                
                flash(f'成功添加 {added_count} 个新关键词！', 'success')
            
            if skipped_count > 0:
                flash(f'跳过了 {skipped_count} 个已存在的关键词。', 'info')

        return redirect(url_for('keywords'))

    # 处理GET请求
    search_query = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)  # 获取当前页码，默认第1页
    per_page = 50  # 每页显示50条
    
    query = Keyword.query
    if search_query:
        query = query.filter(Keyword.text.ilike(f'%{search_query}%'))

    # 使用分页：paginate(page, per_page, error_out=False)
    pagination = query.order_by(Keyword.id.desc()).paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    
    all_groups = MonitoredGroup.query.all()
    return render_template(
        'keywords.html', 
        keywords=pagination.items,  # 当前页的关键词
        pagination=pagination,       # 分页对象
        groups=all_groups, 
        search_query=search_query
    )

@app.route('/keywords/edit/<int:keyword_id>', methods=['GET', 'POST'])
@login_required # 添加鉴权装饰器
def edit_keyword(keyword_id):
    keyword_to_edit = Keyword.query.get_or_404(keyword_id)
    if request.method == 'POST':
        group_ids = request.form.getlist('groups')
        if not group_ids:
            flash('必须至少选择一个群组。', 'danger')
        else:
            # 性能优化: 保存旧的群组列表
            old_groups = list(keyword_to_edit.groups)
            
            groups = MonitoredGroup.query.filter(MonitoredGroup.id.in_(group_ids)).all()
            keyword_to_edit.groups = groups 
            db.session.commit()
            
            # 性能优化: 清空旧的和新的群组的AC自动机缓存
            with automatons_lock:
                for group in old_groups:
                    keyword_automatons.pop(group.id, None)
                for group in groups:
                    keyword_automatons.pop(group.id, None)
            
            flash('关键词关联已更新！', 'success')
        return redirect(url_for('keywords'))

    all_groups = MonitoredGroup.query.all()
    linked_group_ids = {group.id for group in keyword_to_edit.groups}
    return render_template('edit_keyword.html', keyword=keyword_to_edit, groups=all_groups, linked_group_ids=linked_group_ids)


@app.route('/keywords/delete/<int:keyword_id>')
@login_required # 添加鉴权装饰器
def delete_keyword(keyword_id):
    keyword_to_delete = Keyword.query.get_or_404(keyword_id)
    
    # 性能优化: 清空相关群组的AC自动机缓存
    affected_groups = list(keyword_to_delete.groups)
    
    db.session.delete(keyword_to_delete)
    db.session.commit()
    
    # 删除后清空缓存
    with automatons_lock:
        for group in affected_groups:
            keyword_automatons.pop(group.id, None)
    
    flash('关键词已删除。', 'info')
    return redirect(url_for('keywords'))

@app.route('/keywords/batch_delete', methods=['POST'])
@login_required # 添加鉴权装饰器
def batch_delete_keywords():
    keyword_ids = request.form.getlist('keyword_ids')
    if not keyword_ids:
        flash('没有选择任何关键词。', 'warning')
        return redirect(url_for('keywords'))
        
    keywords_to_delete = Keyword.query.filter(Keyword.id.in_(keyword_ids)).all()
    
    # 性能优化: 收集所有受影响的群组
    affected_groups = set()
    for keyword in keywords_to_delete:
        for group in keyword.groups:
            affected_groups.add(group.id)
    
    deleted_count = len(keywords_to_delete)
    for keyword in keywords_to_delete:
        db.session.delete(keyword)
    
    db.session.commit()
    
    # 清空所有受影响群组的AC自动机缓存
    with automatons_lock:
        for group_id in affected_groups:
            keyword_automatons.pop(group_id, None)
    
    flash(f'成功删除 {deleted_count} 个关键词！', 'success')
    return redirect(url_for('keywords'))


@app.route('/messages')
@login_required # 添加鉴权装饰器
def messages():
    group_filter = request.args.get('group_name', '')
    start_date_filter = request.args.get('start_date', '')
    end_date_filter = request.args.get('end_date', '')
    page = request.args.get('page', 1, type=int)  # 获取当前页码，默认第1页
    per_page = 100  # 每页显示100条

    query = MatchedMessage.query

    if group_filter:
        query = query.filter(MatchedMessage.group_name == group_filter)
    if start_date_filter:
        try:
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
            query = query.filter(MatchedMessage.message_date >= start_date)
        except ValueError:
            flash('无效的开始日期格式，请使用 YYYY-MM-DD。', 'danger')
    if end_date_filter:
        try:
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d')
            end_of_day = datetime.combine(end_date, time.max)
            query = query.filter(MatchedMessage.message_date <= end_of_day)
        except ValueError:
            flash('无效的结束日期格式，请使用 YYYY-MM-DD。', 'danger')

    # 使用分页
    pagination = query.order_by(MatchedMessage.message_date.desc()).paginate(
        page=page, 
        per_page=per_page, 
        error_out=False
    )
    
    all_message_groups = db.session.query(MatchedMessage.group_name).distinct().order_by('group_name').all()
    unique_group_names = [name for name, in all_message_groups]

    all_groups = MonitoredGroup.query.all()
    group_logo_map = {g.group_name: g.logo_path for g in all_groups}
    
    filter_values = {
        'group_name': group_filter,
        'start_date': start_date_filter,
        'end_date': end_date_filter
    }
    
    return render_template(
        'messages.html', 
        messages=pagination.items,  # 当前页的消息
        pagination=pagination,       # 分页对象
        group_logo_map=group_logo_map,
        unique_group_names=unique_group_names,
        filter_values=filter_values
    )


@app.route('/messages/export')
@login_required
def export_messages():
    """导出消息为Excel文件，支持中文和特殊字符"""
    try:
        # 获取筛选条件（与messages路由相同）
        group_filter = request.args.get('group_name', '')
        start_date_filter = request.args.get('start_date', '')
        end_date_filter = request.args.get('end_date', '')
        
        # 构建查询（与messages路由相同）
        query = MatchedMessage.query
        
        if group_filter:
            query = query.filter(MatchedMessage.group_name == group_filter)
        if start_date_filter:
            try:
                start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
                query = query.filter(MatchedMessage.message_date >= start_date)
            except ValueError:
                pass
        if end_date_filter:
            try:
                end_date = datetime.strptime(end_date_filter, '%Y-%m-%d')
                end_of_day = datetime.combine(end_date, time.max)
                query = query.filter(MatchedMessage.message_date <= end_of_day)
            except ValueError:
                pass
        
        # 获取所有符合条件的消息（不分页）
        messages = query.order_by(MatchedMessage.message_date.desc()).all()
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "消息记录"
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置表头
        headers = ['序号', '群组名称', '匹配关键词', '发送者', '消息内容', '消息时间']
        ws.append(headers)
        
        # 设置表头样式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8   # 序号
        ws.column_dimensions['B'].width = 25  # 群组名称
        ws.column_dimensions['C'].width = 20  # 匹配关键词
        ws.column_dimensions['D'].width = 20  # 发送者
        ws.column_dimensions['E'].width = 60  # 消息内容
        ws.column_dimensions['F'].width = 20  # 消息时间
        
        # 写入数据
        for idx, msg in enumerate(messages, start=1):
            row_data = [
                idx,
                msg.group_name or '',
                msg.matched_keyword or '',
                msg.sender or '未知',
                msg.message_content or '',
                msg.message_date.strftime('%Y-%m-%d %H:%M:%S') if msg.message_date else ''
            ]
            ws.append(row_data)
            
            # 设置数据行样式
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=idx + 1, column=col_num)
                cell.border = border_style
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                # 交替行颜色
                if idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename_parts = ['消息记录']
        if group_filter:
            filename_parts.append(f'_{group_filter}')
        if start_date_filter:
            filename_parts.append(f'_{start_date_filter}')
        if end_date_filter:
            filename_parts.append(f'至{end_date_filter}')
        
        filename = ''.join(filename_parts) + '.xlsx'
        
        # 返回文件
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename  # Flask 2.0+ 使用 download_name
        )
        
    except Exception as e:
        flash(f'导出失败: {str(e)}', 'danger')
        return redirect(url_for('messages'))

    
@app.route('/messages/delete/<int:message_id>')
@login_required # 添加鉴权装饰器
def delete_message(message_id):
    message_to_delete = MatchedMessage.query.get_or_404(message_id)
    db.session.delete(message_to_delete)
    db.session.commit()
    flash('消息已删除。', 'info')
    return redirect(url_for('messages'))

@app.route('/messages/clear_all')
@login_required # 添加鉴权装饰器
def clear_all_messages():
    try:
        num_rows_deleted = db.session.query(MatchedMessage).delete()
        db.session.commit()
        flash(f'已清空 {num_rows_deleted} 条消息。', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'清空消息时出错: {e}', 'danger')
    return redirect(url_for('messages'))


# WebSocket 事件处理
@socketio.on('connect')
def handle_connect():
    """客户端连接事件"""
    print(f"[WebSocket] 客户端已连接")

@socketio.on('disconnect')
def handle_disconnect():
    """客户端断开连接事件"""
    print(f"[WebSocket] 客户端已断开")

# 广播新消息到所有连接的客户端
def broadcast_new_message(message_data):
    """
    广播新消息到所有WebSocket客户端
    message_data: {
        'group_name': str,
        'sender': str,
        'matched_keyword': str,
        'message_content': str,
        'message_date': str
    }
    """
    try:
        socketio.emit('new_message', message_data, namespace='/')
        print(f"[WebSocket] 广播新消息: {message_data.get('matched_keyword')}")
    except Exception as e:
        print(f"[WebSocket] 广播失败: {e}")

if __name__ == '__main__':
    WEB_HOST = os.environ.get('WEB_HOST', '127.0.0.1')  # 默认绑定本地，仅允许本地访问
    WEB_PORT = int(os.environ.get('WEB_PORT', 8033))

    # --- 数据库初始化 & 管理员创建 --- #
    with app.app_context():
        # 自动升级数据库
        auto_upgrade_database()

        # 创建默认管理员用户
        if User.query.count() == 0:
            default_username = "admin"
            env_admin_pw = os.environ.get('ADMIN_PASSWORD', '')
            default_password = env_admin_pw if env_admin_pw else secrets.token_urlsafe(16)
            hashed_password = generate_password_hash(default_password)

            new_admin_user = User(
                username=default_username,
                password_hash=hashed_password,
                is_admin=True
            )
            db.session.add(new_admin_user)
            db.session.commit()

            print("------------------------------------------------------")
            print("⚠️  初次运行：已创建默认管理员用户！")
            print(f"   用户名: {default_username}")
            print(f"   密码:   {default_password}")
            print("   请务必妥善保管此密码")
            print("------------------------------------------------------")

    # --- 启动 Telegram 监控（如果已配置，异步启动） --- #
    with app.app_context():
        config = Config.query.first()
        if config and config.api_id and config.api_hash and config.phone_number:
            print("[启动] 检测到 Telegram 配置，异步启动监控服务...")
            def start_monitoring_async():
                try:
                    import telegram_monitor
                    telegram_monitor.websocket_broadcast_callback = broadcast_new_message
                    start_monitoring()
                    print("[启动] Telegram 监控服务已就绪！")
                except Exception as e:
                    print(f"[启动] ⚠️ 监控启动失败: {e}")

            thread = Thread(target=start_monitoring_async, daemon=True)
            thread.start()
        else:
            print("[启动] 未配置 Telegram，将通过 Web 界面配置")

    # --- 启动 Web 服务器 --- #
    print(f"\n🚀 Web 服务器启动中 (端口: {WEB_PORT}, 绑定: {WEB_HOST})...")
    print(f"   访问地址: http://localhost:{WEB_PORT}")
    socketio.run(app, host=WEB_HOST, port=WEB_PORT, debug=False, allow_unsafe_werkzeug=True)